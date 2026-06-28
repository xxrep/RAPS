#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generate RAPS_rebuttal_experiments.xlsx from reviews.md.
Baselines/datasets/ablation/robustness are filled from the paper's AUTHORITATIVE text
(2602.08009v1.txt, Tables 1/2/3) — NOT image OCR, NOT invented. Existing values are pre-filled;
only genuinely-new conditions are left yellow (to run). Each sheet has a 可行性/成本 line."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook(); wb.remove(wb.active)
TITLE_FILL = PatternFill("solid", fgColor="1F4E78"); HDR_FILL = PatternFill("solid", fgColor="2E75B6")
META_FILL = PatternFill("solid", fgColor="DDEBF7"); FEAS_FILL = PatternFill("solid", fgColor="FCE4D6")
COND_FILL = PatternFill("solid", fgColor="F2F2F2"); TODO_FILL = PatternFill("solid", fgColor="FFF2CC")
RAPS_FILL = PatternFill("solid", fgColor="E2EFDA")
thin = Side(style="thin", color="BFBFBF"); BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top"); CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)

BENCH = ["MMLU", "GSM8K", "SVAMP", "AQuA", "HumanEval", "平均"]
PAPER_METHODS = [
 ("单智能体", "Vanilla IO (GPT-4o-mini)"), ("单智能体", "CoT"), ("单智能体", "ComplexCoT"), ("单智能体", "SC"),
 ("静态多智能体", "Chain"), ("静态多智能体", "Star"), ("静态多智能体", "Tree"), ("静态多智能体", "Random"),
 ("静态多智能体", "LLM-Debate"), ("静态多智能体", "LLM-Blender"),
 ("通信无关", "GPTSwarm"), ("通信无关", "AgentPrune"), ("通信无关", "AFlow"), ("通信无关", "MaAS"), ("通信无关", "G-Designer"),
 ("元控制", "AutoAgents"), ("元控制", "Puppeteer"), ("元控制", "MAS-Zero"),
 ("动态自组网", "RAPS (Ours)"),
]
# Table 1 (paper text, verified by per-row average): [MMLU, GSM8K, SVAMP, AQuA, HumanEval, Average]
T1 = {
 "Vanilla IO (GPT-4o-mini)": [81.7, 91.6, 87.5, 71.3, 72.8, 81.0],
 "CoT": [83.0, 92.1, 88.4, 74.7, 75.7, 82.8], "ComplexCoT": [83.7, 92.5, 89.2, 76.1, 75.2, 83.3],
 "SC": [82.4, 92.4, 88.9, 76.8, 77.5, 83.6], "Chain": [84.3, 91.7, 82.6, 70.4, 81.3, 82.1],
 "Star": [80.4, 91.9, 88.2, 69.6, 74.5, 80.9], "Tree": [82.4, 90.7, 88.5, 73.9, 72.4, 81.6],
 "Random": [85.6, 92.0, 87.0, 75.1, 78.2, 83.6], "LLM-Debate": [85.0, 92.4, 89.8, 77.3, 82.6, 85.4],
 "LLM-Blender": [81.0, 91.3, 88.3, 76.9, "–", 84.4], "GPTSwarm": [83.7, 92.7, 88.5, 78.2, 88.5, 86.3],
 "AgentPrune": [84.3, 92.3, 89.8, 79.1, 86.8, 86.5], "AFlow": [85.6, 94.1, 90.0, 78.5, 91.0, 87.8],
 "MaAS": [85.0, 91.4, 89.3, 76.2, 87.1, 85.8], "G-Designer": [86.3, 93.2, 90.7, 79.4, 90.2, 88.0],
 "AutoAgents": [82.4, 92.5, 86.7, 75.7, 87.6, 85.0], "Puppeteer": [84.3, 93.3, 89.5, 77.5, 75.3, 84.0],
 "MAS-Zero": [83.0, 92.6, 87.3, 72.9, 83.9, 83.9], "RAPS (Ours)": [88.2, 95.4, 92.2, 82.6, 91.5, 90.0],
}
# Table 2 (robustness, MMLU): [5T0A, 4T1A, 3T2A, 2T3A, 5T5A]
T2 = {
 "Chain": [84.3, 72.5, 50.3, 22.2, 16.3], "Random": [85.6, 81.7, 35.3, 18.3, 45.1],
 "LLM-Debate": [85.0, 78.4, 62.1, 30.7, 47.7], "GPTSwarm": [83.7, 75.2, 55.6, 23.5, 52.9],
 "AFlow": [85.6, 79.7, 52.3, 19.6, 28.8], "G-Designer": [86.3, 80.4, 37.9, 15.0, 49.7],
 "Puppeteer-P": [84.3, 77.8, 65.4, 32.0, 51.6], "Puppeteer-C": [84.3, 13.7, "-", "-", "-"],
 "RAPS w/o BR": [86.9, 83.7, 69.3, 33.3, 53.6], "RAPS": [88.2, 87.6, 84.3, 83.0, 86.3],
}
# Table 3 (ablation): [MMLU, GSM8K, HumanEval]
T3 = {
 "RAPS (Full)": [88.2, 95.4, 91.5], "w/o RS": [85.6, 93.7, 89.3], "w/o BR": [86.9, 94.5, 90.7],
 "w/o Both": [83.7, 92.8, 88.5], "w/ LLM Broker": [89.5, 95.9, 93.7], "w/ Naive Agent Pool": [85.0, 93.5, 90.2],
}


def sheet(name, title, reviewers, motivation, design, feasibility, headers, rows, notes=None, widths=None, todo_from=2):
    ws = wb.create_sheet(name[:31]); ncol = len(headers); last = get_column_letter(ncol); r = 1
    ws.merge_cells(f"A{r}:{last}{r}")
    c = ws.cell(r, 1, title); c.font = Font(color="FFFFFF", bold=True, size=13); c.fill = TITLE_FILL
    c.alignment = Alignment(vertical="center"); ws.row_dimensions[r].height = 26; r += 1
    for label, text, fill in [("对应审稿意见", reviewers, META_FILL), ("实验动机与思想", motivation, META_FILL),
                              ("实验设计", design, META_FILL), ("可行性 / 成本", feasibility, FEAS_FILL)]:
        lc = ws.cell(r, 1, label); lc.font = Font(bold=True); lc.fill = fill; lc.alignment = WRAP
        ws.merge_cells(f"B{r}:{last}{r}")
        tc = ws.cell(r, 2, text); tc.alignment = WRAP; tc.fill = fill
        ws.row_dimensions[r].height = max(28, min(210, 15 * (len(text) // 60 + text.count("\n") + 2))); r += 1
    r += 1; header_row = r
    for j, h in enumerate(headers, 1):
        cc = ws.cell(r, j, h); cc.font = Font(color="FFFFFF", bold=True); cc.fill = HDR_FILL
        cc.alignment = CTR; cc.border = BORDER
    ws.row_dimensions[r].height = 30; r += 1
    for row in rows:
        is_raps = isinstance(row[1] if len(row) > 1 else "", str) and "RAPS" in str(row[0]) + str(row[1] if len(row) > 1 else "")
        for j in range(1, ncol + 1):
            val = row[j - 1] if j - 1 < len(row) else ""
            cc = ws.cell(r, j, val); cc.border = BORDER; cc.alignment = WRAP if j == 1 else CTR
            if j == 1:
                cc.fill = COND_FILL; cc.font = Font(bold=True)
            elif val == "" and j >= todo_from:
                cc.fill = TODO_FILL
            elif is_raps:
                cc.fill = RAPS_FILL
        r += 1
    if notes:
        r += 1; ws.merge_cells(f"A{r}:{last}{r}")
        cc = ws.cell(r, 1, "备注：" + notes); cc.alignment = WRAP; cc.font = Font(italic=True, color="595959")
        ws.row_dimensions[r].height = max(28, 15 * (len(notes) // 80 + 2))
    ws.column_dimensions["A"].width = (widths[0] if widths else 28)
    for j in range(2, ncol + 1):
        ws.column_dimensions[get_column_letter(j)].width = (widths[j - 1] if widths and len(widths) >= j else 12)
    ws.freeze_panes = ws.cell(header_row + 1, 3 if ncol > 8 else 2)
    return ws


# E0
ov = [
 ["E1", "实验配置规范", "R1-1, R1-7, R2-4/5/6", "规范/写作", "高", "低", "待办"],
 ["E2", "跨主干模型泛化(全方法×全数据集)", "R1-12, R1-1", "新实验", "高", "高", "待办"],
 ["E3", "预算匹配 + 强单模型基线", "R1-11, R2-3", "新基线", "高(关键)", "中-高", "待办"],
 ["E4", "主表 + 新增基线 DyLAN/MAS-GPT", "R1-8", "新基线", "高", "中-高", "待办"],
 ["E5", "鲁棒性/威胁模型(扩展 Table 2)", "R1-2", "新实验", "高", "中-高", "待办"],
 ["E6", "机制归因诊断", "R1-3, R1-10", "分析", "高", "中", "待办"],
 ["E7", "长程多轮 / SWE-bench", "R1-5, R2-7, R2-10", "新实验", "高(关键)", "高", "进行中"],
 ["E8", "可扩展性—大规模池", "R2-8", "新实验", "中", "中-高", "待办"],
 ["E9", "动态成员—流失&异构", "R1-9", "新实验", "中", "高", "待办"],
 ["E10", "看门狗可靠性", "R1-7.2, R2-9", "分析", "中", "中", "待办"],
 ["E11", "声誉超参敏感性", "R1-7.3", "分析", "中", "低-中", "待办"],
 ["E12", "消融(=Table 3)+naive/crafted说明", "R2-6+现有", "分析/写作", "中", "低", "数值已有"],
 ["—", "写作类(术语/意义/讨论/表1红蓝脚注)", "R1-4/6,R2-1/2", "写作", "高", "低", "待办"],
 ["—", "重画图1、图5", "R1-13", "图", "中", "中", "待办"],
 ["—", "补代码:baselines+对抗攻击setup", "R1-14", "代码", "中", "中", "待办"],
 ["—", "Data/Code Availability+ML checklist", "政策", "—", "高", "低", "待办"],
]
sheet("E0-总览", "RAPS 返修补充实验总览 (Nature Comm.) — 实验 6/20 截止, 7/1 前 rebuttal",
      "汇总 reviews.md 两位审稿人要求的全部补充实验",
      "每张 Sheet = 一个实验:含动机/思想、设计、可直接填写的结果表。绿色=论文已有值(取自 2602.08009v1.txt Tables 1/2/3,已逐行平均校验);黄色=本次待跑;首列=条件。",
      "时间线:5/18 收意见→6/20 完成实验→7/1 rebuttal+cover letter→7/18 提交。逐条回复需逐字引用原文。",
      "所有基线/基准/消融/鲁棒性数值均来自论文权威文本(非图片OCR、非臆造);只补『新增方法/新增条件』。高成本/可能难做的已在各表标注+降级方案。",
      ["编号", "实验 / 表格 (Sheet)", "对应审稿意见", "类型", "优先级", "预估成本", "状态"],
      ov, notes="最贵/最难:E2(全套换backbone重跑)、E7(repo级docker)、E9(动态成员仿真协议);各表『可行性/成本』行有降级方案。",
      widths=[6, 34, 20, 12, 10, 10, 8])

# E1 config (unchanged content)
cfg_rows = [
 ["主干模型", "GPT-4o-mini-2024-07-18", "同一模型、相同实例数", "Table 1 脚注:除单智能体外均 5 个相同配置智能体"],
 ["智能体数量", "5", "5 (相同)", "Table 1 主结果统一 5 个"],
 ["温度 temperature", "0", "0 (相同)", "确定性;网关非完全确定→同题集比较"],
 ["采样数 n / max_tokens", "1 / 2000", "相同", ""],
 ["system prompt / 角色", "沿用 Zhuge/Liu 标准角色定义", "角色集相同", "Implementation Details 明确"],
 ["工具 tools", "代码执行(PAL/code_verify);可选检索", "相同可用性", "需逐方法核对"],
 ["停止 / 聚合", "共识/max_steps/无路由目标;Final Answerer 聚合", "统一", ""],
 ["语义匹配", "text-embedding-3-small + 余弦;阈值≈0.3-0.7;top-k 1-3;兜底取top-1", "(broker专属)", "Impl. Details / R1-7.1 / R2-1b"],
 ["broker", "嵌入匹配(默认)+ LLM broker(见 Table 3)", "—", "对接 E12"],
 ["看门狗", "LLM(同backbone);输入=问题+已知+待评;连贯/扣题/纠错→YES", "—", "详见 E10"],
 ["声誉超参", "先验 Beta(1,1);λ=0.999;ω=0.05-0.1;δ=0.2;τ=0.5/0.3", "—", "R1-7.3;详见 E11"],
]
sheet("E1-配置规范", "E1 实验配置规范（统一、可复现）",
      "R1-1(未说明“配置相同”的含义);R1-7(实现细节缺失:语义匹配/看门狗/声誉超参/采样/停止聚合);R2-4/5/6(表1配置无细节、图4 naive/crafted 未说明)",
      "审稿人担心提升来自实现选择而非协同机制。需逐项列出所有方法共享的精确配置 + RAPS 专属机制参数。",
      "RAPS 列已据代码与论文 Implementation Details 填真实值;基线列在补实验时核对一致性。亦为 E10/E11 锚点。",
      "成本低:整理+核对,无需新跑。需把此表写入正文/附录(R1-7 五项)。",
      ["配置项", "RAPS 取值", "基线(是否一致)", "说明/依据"], cfg_rows,
      widths=[20, 46, 22, 36], todo_from=99)

# E2 cross-backbone — FULL methods × all datasets (GPT-4o-mini filled; new backbone empty)
e2_headers = ["类别", "方法"] + [f"4o-mini·{b}" for b in BENCH] + [f"新backbone·{b}" for b in BENCH]
e2_rows = [[cat, m] + T1[m] + [""] * 6 for cat, m in PAPER_METHODS]
sheet("E2-跨主干模型", "E2 跨主干模型泛化（全方法 × 全数据集；证明非针对单一模型）",
      "R1-12(结果仅基于 GPT-4o-mini,需证明非狭隘调参);R1-1(增加 LLM 类型配置)",
      "证明 RAPS 的协同增益在不同 backbone 上可复现。GPT-4o-mini 列=论文 Table 1(已填全);新 backbone 列待跑。重点看『RAPS−最佳基线』方向是否一致。",
      "固定 5 智能体、相同协同配置(E1),仅换 backbone。开源 backbone 可本地 vLLM(3×A40);第二专有需 API。",
      "成本高:全 19 方法×新 backbone×5 数据集很贵。降级:新 backbone 至少填 RAPS + 每类最强基线(SC/G-Designer/AFlow/Puppeteer)行,其余留空或标 N/A;先报平均亦可。型号待定(开源+第二专有各1)。",
      e2_headers, e2_rows,
      notes="GPT-4o-mini 全部 19×6 已按论文 Table 1 填好(权威文本);新 backbone 为黄色待跑。",
      widths=[12, 22] + [8] * 12, todo_from=9)

# E3 budget-matched — all 5 datasets, weak rows filled
e3_headers = ["方法/设置", "主干"] + BENCH + ["Prompt tok", "Completion tok", "成本($)", "调用数"]
e3_rows = [
 ["单智能体 1×", "GPT-4o-mini"] + T1["Vanilla IO (GPT-4o-mini)"] + ["", "", "", ""],
 ["SC (弱)", "GPT-4o-mini"] + T1["SC"] + ["", "", "", ""],
 ["G-Designer (最强基线)", "GPT-4o-mini"] + T1["G-Designer"] + ["", "", "", ""],
 ["RAPS (Ours, 5×)", "GPT-4o-mini"] + T1["RAPS (Ours)"] + ["", "", "", ""],
 ["强单模型 单次(同预算)", "待定 GPT-4o/Claude"] + [""] * 6 + ["", "", "", ""],
 ["强单模型 CoT/SC(同预算)", "待定 GPT-4o/Claude"] + [""] * 6 + ["", "", "", ""],
]
sheet("E3-预算匹配基线", "E3 预算匹配 + 强单模型基线（关键基线，全数据集）",
      "R2-3(弱×5 累计 token ≈ 强模型单次,缺同预算强单模型基线);R1-11(报告各方法计算/token 成本)",
      "质疑:RAPS 提升是协同优势还是『弱×5≈强』算力差? 在『同 token 预算』下比 RAPS(弱×5) vs 强单模型(1×)。同预算下不输/胜过强单模型即确立协同价值(反而有利于本方法)。",
      "准确率列已按论文 Table 1 填弱模型/RAPS 值;token/成本待测;把强模型单次/CoT/SC 调到与 RAPS 相当总预算。报告 准确率/token 帕累托。",
      "成本中-高:需 GPT-4o/Claude API(密钥+费用)。降级:若专有不可得→用本地较强开源做强单对照并注明。token/成本统计本身低成本(读日志)。",
      e3_headers, e3_rows,
      notes="弱模型与 RAPS 准确率=论文 Table 1;只待补 token/成本 与 强单模型行。",
      widths=[22, 16] + [7] * 6 + [11, 12, 9, 8], todo_from=3)

# E4 main table + new baselines (FULL Table 1 filled)
e4_rows = [[cat, m] + T1[m] for cat, m in PAPER_METHODS]
e4_rows += [["新增基线", "DyLAN"] + [""] * 6, ["新增基线", "MAS-GPT"] + [""] * 6]
sheet("E4-主表+新增基线", "E4 主结果表（= 论文 Table 1，全方法×全数据集）+ DyLAN/MAS-GPT",
      "R1-8(加两个基线:DyLAN 和 MAS-GPT)",
      "完整主结果表:19 个论文方法 × 5 数据集 + 平均(已按权威文本填全),再补 DyLAN(动态早停层级 MAS)、MAS-GPT(生成式 MAS 编排)两条新基线。",
      "相同 5 数据集、相同 backbone、相同 token 统计下加入两条新基线,与 Table 1 同表比较,token 成本纳入 E3。",
      "成本中-高:需复现 DyLAN/MAS-GPT 开源实现并适配本设置。若某个超期→标 N/A 并在 rebuttal 说明。",
      ["类别", "方法"] + BENCH, e4_rows,
      notes="前 19 行=论文 Table 1 完整数值(权威文本,逐行平均已校验);仅 DyLAN/MAS-GPT 两行黄色待跑。代码同时补这两条(R1-14)。",
      widths=[14, 22] + [10] * 6, todo_from=99)

# E5 robustness (Table 2 filled + new adversary types)
e5_rows = [["A1 误导性推理(论文Table2)", m] + T2[m] + [""] for m in
           ["Chain", "Random", "LLM-Debate", "GPTSwarm", "AFlow", "G-Designer", "Puppeteer-P", "Puppeteer-C", "RAPS w/o BR", "RAPS"]]
for a in ["A2 声誉自适应(感知并规避门控)", "A3 伪装有用(先建声誉再投毒)", "A4 利用难本地验证任务"]:
    e5_rows.append([a, "RAPS (w/ BR)", "", "", "", "", "", ""])
    e5_rows.append(["", "RAPS (w/o BR)", "", "", "", "", "", ""])
sheet("E5-鲁棒性威胁模型", "E5 鲁棒性 / 更严谨威胁模型（扩展论文 Table 2，MMLU）",
      "R1-2(鲁棒性过于程式化:需描述威胁模型——对抗提示、是否适应声誉、是否伪装、是否利用难验证任务)",
      "现有 Table 2 仅 A1(被提示传播误导推理)。补更真实威胁:A2 自适应(感知声誉规避)、A3 伪装(先建声誉再投毒)、A4 利用难本地验证任务;并报对抗者检测 TP/FP。",
      "沿用 Table 2 组成(5T0A…5T5A)与 MMLU。A1 行=论文已有值(已填,作对照);A2-A4 跑 RAPS(w/ BR) vs (w/o BR)。每类对抗者的提示模板与能力假设写入附录。",
      "成本中-高:A1 已有;A2/A3(自适应/伪装)设计较复杂(需感知/先建声誉),effort 中-高。若超期→至少补 A4 并定性讨论 A2/A3。",
      ["对抗者类型", "方法", "5T0A", "4T1A", "3T2A", "2T3A", "5T5A", "检测(TP/FP)"], e5_rows,
      notes="组成/基准与论文 Table 2 一致;A1 数值=权威文本。新增仅 A2-A4 × RAPS(w/ vs w/o BR)。",
      widths=[26, 14, 8, 8, 8, 8, 8, 14])

# E6 diagnostics
e6_rows = [
 ["反应式订阅-更新频率", "每题订阅被改写的轮次占比", "", "", "", ""],
 ["反应式订阅-更新幅度", "改写前后订阅文本语义距离", "", "", "", ""],
 ["反应式订阅→路由质量", "更新前/后 broker 语义匹配分", "", "", "", ""],
 ["反应式订阅→下游成功率", "更新后该路由分支正确率", "", "", "", ""],
 ["路由分布", "语义匹配分分布(是否双峰)", "", "", "", ""],
 ["路由图对比", "RAPS vs 静态 vs 控制器 随时间交互图", "", "", "", "代表性任务图"],
 ["声誉↔正确性相关", "声誉分与客观正确性相关系数", "", "", "", ""],
 ["对抗者隔离 FP/FN", "误隔离好agent/漏隔离坏agent 比率", "", "", "", "接 E5"],
 ["不确定性门控影响", "开/关 Beta 浓度门控差异", "", "", "", ""],
]
sheet("E6-机制归因诊断", "E6 机制归因诊断（路由/订阅/声誉的直接证据）",
      "R1-3(机制归因需更直接证据);R1-10(对路由/订阅演变/声誉做机制分析)",
      "『类 MoE 路由、角色可塑性』需超越端到端准确率的机制证据。用直接诊断量证明每个机制按设计起作用。",
      "在主结果同设置下加埋点记录:订阅更新频率/幅度、更新前后匹配分与下游成功率、声誉-正确性相关、隔离 FP/FN、不确定性门控影响;画代表性任务路由图(RAPS vs 静态 vs 控制器)。",
      "成本中:给现有运行加埋点+统计,不新增协同方法;3 个代表基准即可,可扩到全部 5 个。",
      ["机制/诊断指标", "定义/测量", "MMLU", "GSM8K", "HumanEval", "结论/说明"], e6_rows,
      widths=[24, 30, 10, 10, 11, 22])

# E7 SWE-bench
e7_rows = [
 ["单智能体 (mini-swe-agent)", "2/20 (10.0%)", "22.9", "≈250k", "≈0.039", "参照基线(已完成)"],
 ["RAPS (Ours)", "", "", "", "", "进行中(同 20 题固定子集)"],
 ["DyLAN / MAS-GPT (可选)", "", "", "", "", "若时间允许"],
]
sheet("E7-SWEbench长程", "E7 长程多轮 / SWE-bench（repo 级、多步、工具依赖）",
      "R1-5(缺真实多步/富工具/长程协作任务);R2-7(建议评测 SWE-bench);R2-10(更复杂多轮/长程/工具依赖场景)",
      "现有 5 基准多为单轮 I/O,难体现『反应式订阅=动态角色演化』这类本质多步现象。SWE-bench(repo级、多步、跨轮上下文+工具)是检验 RAPS 协同价值的更有意义场景。",
      "SWE-bench Lite 固定子集上,RAPS=多个自主 agent-loop(定位/复现/编辑/验证/评审)经发布订阅+broker 协调,与单智能体(mini-swe-agent)对比 resolved%/轮数/token/成本。统一 gpt-4o-mini。",
      "成本高:repo 级 docker 评测重。已建独立框架(RAPS/swe/),单 agent 基线已完成(2/20);RAPS 调优进行中(诚实:gpt-4o-mini 修复力是瓶颈)。",
      ["方法", "SWE-bench Lite resolved%", "平均轮数", "每题 token", "每题成本($)", "备注"], e7_rows,
      notes="同时回应两审稿人对『真实多步 agentic 任务』的诉求,是稿件意义关键补强。",
      widths=[24, 22, 10, 12, 12, 26])

# E8 scalability
e8_rows = [[str(n), "(图3已有)" if n in (5, 10, 20) else "", "", "", "", "", ""] for n in [5, 10, 20, 40, 80]]
sheet("E8-可扩展性大规模", "E8 可扩展性 — 扩展到更大 agent 池",
      "R2-8(框架强调开放式群体/大规模群体智能,但扩展曲线只到小-中规模)",
      "扩到更大池检验:(i) broker 语义匹配成本是否平滑扩展;(ii) 每 agent 声誉表维护是否可控;(iii) 小规模正向趋势是否持续或饱和。",
      "MMLU 上把智能体数扩到 40/80,记录准确率、broker 匹配耗时、声誉表维护开销、端到端时间、token;5/10/20 复用图 3。",
      "成本中-高:80 较贵。降级:80 可选/粗评,重点给『成本曲线趋势』。",
      ["智能体数量", "准确率(MMLU)", "broker匹配耗时", "声誉表维护", "端到端时间", "总token", "是否饱和"], e8_rows,
      widths=[14, 13, 14, 12, 12, 11, 10], todo_from=2)

# E9 dynamic membership
e9_rows = [
 ["静态拓扑对照", "固定成员(无加入/离开)", "", "", "", "对照基准"],
 ["节点流失-低", "运行中~10% 加入/离开", "", "", "", ""],
 ["节点流失-中", "~30% 中途变动", "", "", "", ""],
 ["节点流失-高", "~50% 中途变动", "", "", "", ""],
 ["异构-混合模型大小", "大/小模型混合", "", "", "", ""],
 ["异构-混合工具访问", "部分有工具/检索,部分没有", "", "", "", ""],
]
sheet("E9-动态成员", "E9 动态成员 — 节点流失 & 异构（验证开放成员资格）",
      "R1-9(成员加入/离开主要是动机陈述而非实测;增加节点流失与异构,验证 §4.1.3)",
      "『开放成员资格/动态发现』未被实证。模拟节点流失(运行中加入/离开)与异构(混合模型大小/工具),验证 RAPS 无需重配拓扑即可适配。",
      "(a)流失:不同比例运行中加入/离开,比 RAPS 与静态拓扑准确率/恢复性;(b)异构:混合模型/工具,看 broker 路由与声誉是否仍有效。",
      "成本高:需新增『运行中加入/离开』仿真协议(本组最不平凡)。降级:先做异构(较易);流失若超期→小规模 demo+定性论证。",
      ["设置", "条件/变量", "准确率", "对静态增益", "稳定性/恢复", "说明"], e9_rows,
      notes="动态成员的仿真协议需在附录写清(审稿人与作者共同关注的难点)。",
      widths=[18, 28, 10, 12, 12, 20])

# E10 watchdog
e10_rows = [
 ["判定准确率", "看门狗 YES/NO 与真值一致率", "", "", ""],
 ["精确率/召回率", "识别『坏发布』的 P/R", "", "", ""],
 ["校准", "判定置信与实际正确吻合(可靠图)", "", "", ""],
 ["对声誉的影响", "误判→声誉误更新的传导", "", "", ""],
 ["实现说明", "是否LLM/看到什么/判定标准", "见E1", "见E1", "需正文澄清"],
]
sheet("E10-看门狗可靠性", "E10 看门狗（watchdog）可靠性评估",
      "R1-7.2(看门狗如何做第一手验证);R2-9(判定如何产生/校准/验证,尤其本身是 LLM 评估器)",
      "贝叶斯声誉直接依赖看门狗判定,但其可靠性未独立评估。需评估判定准确率/校准并澄清实现。",
      "构造带真值标签的发布(有用/误导)集合,测准确率、P/R、校准曲线;分析误判对声誉的传导;正文写清输入与判定标准。",
      "成本中:需小规模标注集(可用对抗实验日志半自动构造)。",
      ["指标", "定义", "GPT-4o-mini", "第二backbone", "说明"], e10_rows,
      widths=[20, 30, 14, 14, 20])

# E11 reputation sensitivity
e11_rows = [
 ["时间衰减 λ", "0.9/0.99/0.999/1.0", "", "", ""],
 ["二手合并权重 ω", "0.0/0.05/0.1/0.2", "", "", ""],
 ["偏离阈值 δ", "0.1/0.2/0.3", "", "", ""],
 ["信任阈值 τ", "0.3/0.5/0.7", "", "", ""],
 ["先验 Beta(α0,β0)", "Beta(1,1)/Beta(2,2)", "", "", ""],
]
sheet("E11-声誉超参敏感性", "E11 声誉超参数敏感性",
      "R1-7.3(声誉超参:衰减λ/招募/加权/不确定性,及敏感性)",
      "审稿人要求报告超参及敏感性,说明结论不依赖精挑细选(鲁棒性来自机制而非调参)。",
      "逐一扫 λ/ω/δ/τ/先验,在干净池(准确率)与对抗池(鲁棒性)各一条曲线;给推荐区间。",
      "成本低-中:单变量扫描,1-2 个基准即可。",
      ["超参数", "扫描取值", "准确率(干净池)", "鲁棒性(对抗池)", "结论"], e11_rows,
      widths=[20, 24, 14, 14, 16])

# E12 ablation (= Table 3, filled)
e12_rows = [[k] + T3[k] + [d] for k, d in [
 ("RAPS (Full)", "完整(论文Table3)"), ("w/o RS", "去反应式订阅"), ("w/o BR", "去贝叶斯声誉"),
 ("w/o Both", "两者皆去"), ("w/ LLM Broker", "broker 用 LLM 而非嵌入"), ("w/ Naive Agent Pool", "通用角色(对比Crafted)")]]
sheet("E12-消融", "E12 消融（= 论文 Table 3）+ naive/crafted 设置说明",
      "现有消融 Table 3;R2-6(图4 naive/crafted 如何设置需说明)",
      "本表数值即论文 Table 3(已填全,无需重跑)。本次=补写 naive(通用角色) vs crafted(专家角色,默认) 池的精确设置与构造方式。",
      "数值取自论文 Table 3(MMLU/GSM8K/HumanEval)。需正文/附录补 Crafted/Naive 池的角色定义、提示、构造规则(R2-6)。",
      "成本低:数值已有;主要写清 naive/crafted 设置(写作)。",
      ["变体", "MMLU", "GSM8K", "HumanEval", "说明"], e12_rows,
      notes="数值=论文 Table 3,勿改;新增工作是 naive/crafted 设置说明。",
      widths=[24, 10, 10, 12, 26], todo_from=99)

out = "/home/lirui/raps_swe/RAPS-main/RAPS_rebuttal_experiments.xlsx"
wb.save(out)
print("saved:", out, "| sheets:", len(wb.sheetnames))
